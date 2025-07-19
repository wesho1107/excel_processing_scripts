import argparse
from typing import Dict, List, Optional, Tuple, Union
import json
import os
import sys

import openpyxl
import pandas as pd


class IdBasedDataDictionaryProcessor:
    """"Process an Excel file data dictionary that uses ID-based mapping between tables and columns."""
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        
    def load_excel(self) -> bool:
        """Load Excel file and select worksheet."""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            return True
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return False
        
    def convert_all_sheets_to_markdown(self, output_dir: str) -> List[str]:
        """
        Convert all sheets in the Excel file to markdown files for AI analysis.
        
        Args:
            output_dir: Directory to save markdown files
            
        Returns:
            List of generated file paths
        """
        if not self.load_excel():
            return []
        
        os.makedirs(output_dir, exist_ok=True)
        generated_files = []
        
        for sheet_name in self.workbook.sheetnames:
            print(f"Processing sheet: {sheet_name}")
            
            worksheet = self.workbook[sheet_name]
            markdown_content = self.sheet_to_markdown(worksheet, sheet_name)
            
            # Create safe filename
            safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            output_file = os.path.join(output_dir, f"{safe_name}.md")
            
            try:
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(markdown_content)
                print(f"Saved {sheet_name} to {output_file}")
                generated_files.append(output_file)
            except Exception as e:
                print(f"Error saving {sheet_name}: {e}")
        
        return generated_files
    
    def sheet_to_markdown(self, worksheet, sheet_name: str) -> str:
        """Convert an entire worksheet to markdown format with cell references for AI analysis."""
        lines = [f"# Sheet: {sheet_name}\n", "## Raw Data with Cell References\n"]
        
        # Get dimensions
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row == 0 or max_col == 0:
            lines.extend([
                "No data found in worksheet.",
                "\n## Summary for AI Analysis",
                f"- Total rows with data: 0",
                f"- Total columns: 0",
                f"- Sheet name: {sheet_name}"
            ])
            return "\n".join(lines)
        
        # Pre-calculate column letters for efficiency
        col_letters = [openpyxl.utils.get_column_letter(col) for col in range(1, max_col + 1)]
        
        # Always use the optimized iter_rows approach (works for all sizes)
        data_rows = []
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_row, 
                                                        min_col=1, max_col=max_col, 
                                                        values_only=True), 1):
            row_data = []
            has_content = False
            
            for col_idx, cell_value in enumerate(row):
                if cell_value is not None:
                    col_letter = col_letters[col_idx]
                    row_data.append(f"{cell_value} ({col_letter}{row_idx})")
                    has_content = True
                else:
                    row_data.append("")
            
            if has_content:
                data_rows.append(row_data)
        
        # Convert to DataFrame and generate markdown table
        if data_rows:
            df = pd.DataFrame(data_rows, columns=[f"Col {i+1}" for i in range(max_col)])
            lines.append(df.to_markdown(index=False, tablefmt='github'))
        else:
            lines.append("No data found in worksheet.")
        
        # Add summary
        lines.extend([
            "\n## Summary for AI Analysis",
            f"- Total rows with data: {len(data_rows)}",
            f"- Total columns: {max_col}",
            f"- Sheet name: {sheet_name}",
            "- Cell references are included in parentheses for precise identification"
        ])
        
        return "\n".join(lines)
    
    def filter_data_by_criteria(self, start_cell: str, columns: Optional[List[str]], 
                            primary_column: str, key_value_mapping: Dict[str, Union[str, int]], 
                            output_path: str, sheet_name: Optional[str] = None) -> Dict[str, str]:
        """
        Filter data for single or multiple values using a single DataFrame load for optimal performance.
        
        Args:
            start_cell: Cell reference where data table starts (e.g., "A6")
            columns: List of column names to extract (None = all columns)
            primary_column: Column to filter by
            key_value_mapping: Dictionary mapping output names to filter values 
            output_path: Output file path (single .md file) or directory (for multiple files)
            sheet_name: Sheet name to process (optional)
            
        Returns:
            Dictionary mapping output names to generated markdown content
        """
        print(f"Starting data extraction from cell {start_cell}")
        print(f"Processing {len(key_value_mapping)} filter value(s)")
        
        # Load data into DataFrame once
        df = self._load_data_as_dataframe(start_cell, sheet_name)
        if df is None:
            print("Failed to load data into DataFrame")
            return {}
        
        print(f"Loaded DataFrame with shape: {df.shape}")
        print(f"Available columns: {list(df.columns)}")
        
        # Validate primary column exists
        if primary_column not in df.columns:
            print(f"Error: Primary column '{primary_column}' not found in data.")
            print(f"Available columns: {list(df.columns)}")
            return {}
        
        # If no columns specified, use all available columns
        if columns is None:
            columns = list(df.columns)
            print(f"No columns specified, using all columns: {columns}")
        else:
            # Validate requested columns exist
            missing_columns = [col for col in columns if col not in df.columns]
            if missing_columns:
                print(f"Warning: Columns not found: {missing_columns}")
                columns = [col for col in columns if col in df.columns]
            
            if not columns:
                print("Error: None of the specified columns were found.")
                return {}
        
        # Process each filter value
        results = {}
        successful_filters = 0
        
        for output_name, key_value in key_value_mapping.items():
            print(f"Processing {output_name}: {primary_column} = '{key_value}'")
            
            # Filter data using pandas
            filtered_df = df[df[primary_column].astype(str).str.strip().str.lower() == str(key_value).lower()]
            
            if filtered_df.empty:
                print(f"  No data found for {primary_column} = '{key_value}'")
                continue
            
            # Select only requested columns
            filtered_df = filtered_df[columns].copy()
            
            print(f"  Found {len(filtered_df)} rows matching criteria")
            
            # Create markdown content
            markdown_content = self._create_filtered_markdown(filtered_df, primary_column, str(key_value), start_cell)
            
            file_path = os.path.join(output_path, f"{output_name}.md")
            
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(markdown_content)
                print(f"  Saved to: {file_path}")
                results[output_name] = markdown_content
                successful_filters += 1
            except Exception as e:
                print(f"  Error saving file: {e}")
        
        print(f"Processing completed: {successful_filters}/{len(key_value_mapping)} filter(s) processed successfully")
        return results

    def _load_data_as_dataframe(self, start_cell: str, sheet_name: Optional[str] = None) -> Optional[pd.DataFrame]:
        """
        Load Excel data into DataFrame.
        
        Args:
            start_cell: Cell reference where data table starts (e.g., "A6")
            sheet_name: Sheet name to process (optional)
            
        Returns:
            DataFrame with the data or None if failed
        """
        try:
            # Parse start cell to get row and column
            start_row, start_col = self._parse_cell_reference(start_cell)
            
            # Use pandas to read Excel directly
            if sheet_name is None:
                df = pd.read_excel(
                    self.file_path,
                    header=start_row - 1,  # pandas uses 0-based indexing
                    usecols=lambda x: x >= start_col - 1 if isinstance(x, int) else True,
                    engine='openpyxl'
                )
            else:
                df = pd.read_excel(
                    self.file_path,
                    sheet_name=sheet_name,
                    header=start_row - 1,  # pandas uses 0-based indexing
                    usecols=lambda x: x >= start_col - 1 if isinstance(x, int) else True,
                    engine='openpyxl'
                )
            
            # Clean column names
            df.columns = df.columns.astype(str).str.strip()
            
            # Remove completely empty rows and columns
            df = df.dropna(how='all', axis=0).dropna(how='all', axis=1)
            
            return df
            
        except Exception as e:
            print(f"Error loading data as DataFrame: {e}")
            return None
        
    def _parse_cell_reference(self, cell_ref: str) -> Tuple[int, int]:
        """Parse cell reference like 'A3' into row and column numbers."""
        col_str = ""
        row_str = ""
        
        for char in cell_ref:
            if char.isalpha():
                col_str += char
            else:
                row_str += char
        
        col_num = openpyxl.utils.column_index_from_string(col_str)
        row_num = int(row_str)
        
        return row_num, col_num
    
    def _create_filtered_markdown(self, df: pd.DataFrame, primary_column: str, 
                                 key_value: str, start_cell: str) -> str:
        """Create markdown content with metadata for filtered data."""
        lines = [
            f"# Filtered Data: {primary_column} = '{key_value}'",
            "",
            "## Metadata",
            f"- Filter criteria: {primary_column} = '{key_value}'",
            f"- Data source starting cell: {start_cell}",
            f"- Total matching rows: {len(df)}",
            f"- Columns included: {', '.join(df.columns)}",
            "",
            "## Data Table",
            ""
        ]
        
        # Add the actual data table
        lines.append(df.to_markdown(index=False))
        
        return "\n".join(lines)

def main():
    """Main function for command-line interface with AI-assisted workflow support."""
    parser = argparse.ArgumentParser(
        description="ID-Based Data Dictionary Processor - AI-Assisted Workflow",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
AI-Assisted Workflow Examples:

Convert all sheets to markdown for AI analysis:
  python id_based_data_dictionary_processor.py input.xlsx --convert-all

Batch filter with custom mapping (JSON file):
  python id_based_data_dictionary_processor.py input.xlsx -s "A6" -c "UserID" "Name" "Age" -pc "UserID" --key-value-mapping "user_mapping.json" -o "output_dir/"

Note: JSON mapping file should contain: {"table_1": "table_id_1", "table_2": "table_id_2", ...}
        """
    )
    
    parser.add_argument("input_file", help="Path to the Excel file")
    
    # AI-assisted workflow options
    parser.add_argument("--convert-all", action="store_true",
                       help="Convert all sheets to markdown files for AI analysis")
    parser.add_argument("-s", "--start-cell", 
                       help="Starting cell of data table (e.g., 'A6') for filtering")
    parser.add_argument("-c", "--columns", nargs="*",
                       help="List of columns to extract. If not specified, all columns will be included.")
    parser.add_argument("-pc", "--primary-column", 
                       help="Primary column to filter")
    parser.add_argument("-k", "--key-value-mapping", 
                       help="JSON file path containing key-value mapping to filter for in primary column")
    parser.add_argument("-o", "--output", help="Output directory for generated files")
    parser.add_argument("--sheet", help="Sheet name to process")
    
    args = parser.parse_args()
    
    # Validate input file
    if not os.path.exists(args.input_file):
        print(f"Error: Input file '{args.input_file}' does not exist.")
        sys.exit(1)
    
    processor = IdBasedDataDictionaryProcessor(args.input_file)
    
    # Convert all sheets to markdown
    if args.convert_all:
        print("Converting all sheets to markdown for AI analysis...")
        output_dir = args.output or "sheets_markdown"
        generated_files = processor.convert_all_sheets_to_markdown(output_dir)
        
        if generated_files:
            print(f"\nGenerated {len(generated_files)} markdown files:")
            for file in generated_files:
                print(f"  - {file}")
            print(f"\nNext: Have the AI analyze these files to identify:")
            print("  - Table and column ID mappings")
            print("  - Starting cell of actual data, where the headers begins")
            print("  - Mapping keys and relationships")
        else:
            print("Failed to generate markdown files.")
            sys.exit(1)
        return
    
    # Check for filtering parameters
    if not args.start_cell or not args.primary_column:
        print("Error: Filtering requires --start-cell and --primary-column.")
        sys.exit(1)
    
    if not args.output:
        print("Error: Output path (-o) is required for filtering.")
        sys.exit(1)
    
    if not args.key_value_mapping:
        print("Error: JSON mapping file (-k) is required for filtering.")
        sys.exit(1)
    
    # Load mapping from JSON file
    try:
        with open(args.key_value_mapping, 'r', encoding='utf-8') as f:
            key_value_mapping = json.load(f)
        print(f"Loaded {len(key_value_mapping)} mappings from {args.key_value_mapping}")
    except Exception as e:
        print(f"Error loading mapping file: {e}")
        sys.exit(1)
    
    if not key_value_mapping:
        print("Error: JSON mapping file is empty.")
        sys.exit(1)
    
    # Create output directory
    os.makedirs(args.output, exist_ok=True)
    
    # Execute filtering
    print("Starting data filtering...")
    results = processor.filter_data_by_criteria(
        args.start_cell, args.columns, args.primary_column, 
        key_value_mapping, args.output, args.sheet
    )
    
    if not results:
        print("Filtering failed - no data was processed.")
        sys.exit(1)
    
    print(f"Filtering completed successfully: {len(results)} file(s) generated.")


if __name__ == "__main__":
    main()