#!/usr/bin/env python3
"""
Create a sample Excel file for testing the Excel processor.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook

def create_sample_excel():
    """Create a sample Excel file with some title cells and a data table."""
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Data"
    
    # Add some title/header cells before the actual data
    ws['A1'] = "Company Report"
    ws['A2'] = "Employee Data"
    ws['A3'] = "Generated on: July 15, 2025"
    
    # Add some empty rows
    # Row 4 is empty
    # Row 5 is empty
    
    # Add column headers starting at row 6
    headers = ['Name', 'Age', 'Job', 'Department', 'Salary', 'Birthday']
    for col, header in enumerate(headers, 1):
        ws.cell(row=6, column=col, value=header)
    
    # Add sample data
    sample_data = [
        ['John Doe', 30, 'Software Engineer', 'IT', 75000, '1993-05-15'],
        ['Jane Smith', 28, 'Data Analyst', 'Analytics', 65000, '1995-08-22'],
        ['Bob Johnson', 35, 'Product Manager', 'Product', 85000, '1988-12-10'],
        ['Alice Brown', 32, 'UX Designer', 'Design', 70000, '1991-03-07'],
        ['Charlie Wilson', 29, 'DevOps Engineer', 'IT', 72000, '1994-11-18'],
        ['Diana Lee', 31, 'Marketing Manager', 'Marketing', 68000, '1992-06-25'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 7):  # Start from row 7
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Save the file
    wb.save("sample_data.xlsx")
    print("Created sample_data.xlsx with test data")
    
    # Also create a more complex example with multiple sheets
    wb2 = Workbook()
    
    # Remove default sheet
    wb2.remove(wb2.active)
    
    # Create Sheet1 with employee data
    ws1 = wb2.create_sheet("Employees")
    ws1['B2'] = "Employee Information"
    ws1['B3'] = "Department: Engineering"
    
    # Headers in row 5
    emp_headers = ['Employee ID', 'Name', 'Job', 'Email', 'Phone']
    for col, header in enumerate(emp_headers, 2):  # Start from column B
        ws1.cell(row=5, column=col, value=header)
    
    emp_data = [
        ['E001', 'John Doe', 'Software Engineer', 'john@company.com', '555-0101'],
        ['E002', 'Jane Smith', 'Data Analyst', 'jane@company.com', '555-0102'],
        ['E003', 'Bob Johnson', 'Product Manager', 'bob@company.com', '555-0103'],
    ]
    
    for row_idx, row_data in enumerate(emp_data, 6):
        for col_idx, cell_value in enumerate(row_data, 2):
            ws1.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Create Sheet2 with project data
    ws2 = wb2.create_sheet("Projects")
    ws2['A1'] = "Project Status Report"
    
    proj_headers = ['Project Name', 'Manager', 'Status', 'Deadline']
    for col, header in enumerate(proj_headers, 1):
        ws2.cell(row=3, column=col, value=header)
    
    proj_data = [
        ['Website Redesign', 'John Doe', 'In Progress', '2025-08-01'],
        ['Mobile App', 'Jane Smith', 'Planning', '2025-09-15'],
        ['Data Migration', 'Bob Johnson', 'Completed', '2025-07-01'],
    ]
    
    for row_idx, row_data in enumerate(proj_data, 4):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=cell_value)
    
    wb2.save("complex_sample.xlsx")
    print("Created complex_sample.xlsx with multiple sheets")

if __name__ == "__main__":
    create_sample_excel()
