#!/usr/bin/env python3
"""
Advanced Excel CLI Processor - Converted from Excel MCP Server

A comprehensive command-line tool for advanced Excel operations including:
- Data manipulation and analysis
- Formatting and styling
- Chart creation
- Pivot table generation
- Formula operations
- Worksheet management

Based on the excel-mcp-server by haris-musa
"""

import argparse
import sys
import os
import json
from pathlib import Path
from typing import List, Optional, Dict, Any
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, AreaChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, column_index_from_string
import uuid


class AdvancedExcelProcessor:
    """Advanced Excel processor with comprehensive Excel manipulation capabilities."""
    
    def __init__(self, file_path: str):
        """Initialize the advanced Excel processor."""
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def load_excel(self, sheet_name: Optional[str] = None) -> bool:
        """Load Excel file and select worksheet."""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            if sheet_name:
                if sheet_name in self.workbook.sheetnames:
                    self.worksheet = self.workbook[sheet_name]
                else:
                    print(f"Warning: Sheet '{sheet_name}' not found. Using first sheet.")
                    self.worksheet = self.workbook.active
            else:
                self.worksheet = self.workbook.active
            return True
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return False
    
    def create_workbook(self, filepath: str) -> bool:
        """Create a new Excel workbook."""
        try:
            wb = openpyxl.Workbook()
            wb.save(filepath)
            print(f"Created new workbook: {filepath}")
            return True
        except Exception as e:
            print(f"Error creating workbook: {e}")
            return False
    
    def get_workbook_metadata(self, include_ranges: bool = False) -> Dict[str, Any]:
        """Get comprehensive metadata about the workbook."""
        if not self.load_excel():
            return {}
        
        metadata = {
            "filename": os.path.basename(self.file_path),
            "sheets": []
        }
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_info = {
                "name": sheet_name,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "data_range": f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            }
            
            if include_ranges:
                # Add more detailed range information
                sheet_info["merged_cells"] = [str(range_) for range_ in sheet.merged_cells.ranges]
                sheet_info["has_data"] = sheet.max_row > 1 or sheet.max_column > 1
            
            metadata["sheets"].append(sheet_info)
        
        return metadata
    
    def read_data_range(self, sheet_name: str, start_cell: str = "A1", 
                       end_cell: Optional[str] = None, preview_only: bool = False) -> List[List[Any]]:
        """Read data from a specific range."""
        if not self.load_excel(sheet_name):
            return []
        
        if end_cell:
            range_string = f"{start_cell}:{end_cell}"
        else:
            range_string = f"{start_cell}:{get_column_letter(self.worksheet.max_column)}{self.worksheet.max_row}"
        
        data = []
        for row in self.worksheet[range_string]:
            row_data = [cell.value for cell in row]
            data.append(row_data)
            if preview_only and len(data) >= 10:  # Limit preview to 10 rows
                break
        
        return data
    
    def write_data_to_range(self, sheet_name: str, data: List[List[Any]], 
                           start_cell: str = "A1") -> bool:
        """Write data to a specific range."""
        if not self.load_excel(sheet_name):
            return False
        
        try:
            start_row = int(''.join(filter(str.isdigit, start_cell)))
            start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
            
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    self.worksheet.cell(row=start_row + row_idx, 
                                      column=start_col + col_idx, 
                                      value=value)
            
            self.workbook.save(self.file_path)
            print(f"Data written to {sheet_name} starting at {start_cell}")
            return True
        except Exception as e:
            print(f"Error writing data: {e}")
            return False
    
    def format_range(self, sheet_name: str, start_cell: str, end_cell: str,
                    bold: bool = False, italic: bool = False, 
                    font_size: Optional[int] = None, font_color: Optional[str] = None,
                    bg_color: Optional[str] = None, border_style: Optional[str] = None) -> bool:
        """Apply formatting to a range of cells."""
        if not self.load_excel(sheet_name):
            return False
        
        try:
            # Create font with specified properties
            font_kwargs = {}
            if bold:
                font_kwargs['bold'] = True
            if italic:
                font_kwargs['italic'] = True
            if font_size:
                font_kwargs['size'] = font_size
            if font_color:
                font_kwargs['color'] = font_color
            
            font = Font(**font_kwargs) if font_kwargs else None
            
            # Create fill if background color specified
            fill = PatternFill(start_color=bg_color, end_color=bg_color, 
                             fill_type='solid') if bg_color else None
            
            # Create border if specified
            border = None
            if border_style:
                side = Side(style=border_style)
                border = Border(left=side, right=side, top=side, bottom=side)
            
            # Apply formatting to range
            for row in self.worksheet[f"{start_cell}:{end_cell}"]:
                for cell in row:
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if border:
                        cell.border = border
            
            self.workbook.save(self.file_path)
            print(f"Formatting applied to range {start_cell}:{end_cell}")
            return True
        except Exception as e:
            print(f"Error applying formatting: {e}")
            return False
    
    def create_chart(self, sheet_name: str, data_range: str, chart_type: str,
                    target_cell: str, title: str = "", x_axis: str = "", y_axis: str = "") -> bool:
        """Create a chart in the worksheet."""
        if not self.load_excel(sheet_name):
            return False
        
        try:
            # Create chart based on type
            if chart_type.lower() == 'line':
                chart = LineChart()
            elif chart_type.lower() == 'bar':
                chart = BarChart()
            elif chart_type.lower() == 'pie':
                chart = PieChart()
            elif chart_type.lower() == 'scatter':
                chart = ScatterChart()
            elif chart_type.lower() == 'area':
                chart = AreaChart()
            else:
                print(f"Unsupported chart type: {chart_type}")
                return False
            
            # Set chart properties
            chart.title = title
            chart.x_axis.title = x_axis
            chart.y_axis.title = y_axis
            
            # Add data to chart
            data = Reference(self.worksheet, range_string=data_range)
            chart.add_data(data, titles_from_data=True)
            
            # Add chart to worksheet
            self.worksheet.add_chart(chart, target_cell)
            
            self.workbook.save(self.file_path)
            print(f"Chart created at {target_cell}")
            return True
        except Exception as e:
            print(f"Error creating chart: {e}")
            return False
    
    def create_pivot_table(self, sheet_name: str, data_range: str, 
                          rows: List[str], values: List[str], 
                          columns: Optional[List[str]] = None,
                          agg_func: str = "sum") -> bool:
        """Create a pivot table (simplified version)."""
        if not self.load_excel(sheet_name):
            return False
        
        try:
            # Read source data
            source_data = self.read_data_range(sheet_name, data_range.split(':')[0], data_range.split(':')[1])
            if not source_data:
                print("No data found in specified range")
                return False
            
            # Convert to DataFrame for easier manipulation
            headers = source_data[0]
            data_rows = source_data[1:]
            df = pd.DataFrame(data_rows, columns=headers)
            
            # Create pivot table
            pivot_df = df.pivot_table(
                index=rows,
                values=values,
                columns=columns,
                aggfunc=agg_func,
                fill_value=0
            )
            
            # Create new sheet for pivot table
            pivot_sheet_name = f"{sheet_name}_pivot"
            if pivot_sheet_name in self.workbook.sheetnames:
                self.workbook.remove(self.workbook[pivot_sheet_name])
            
            pivot_sheet = self.workbook.create_sheet(pivot_sheet_name)
            
            # Write pivot table to new sheet
            pivot_data = [pivot_df.columns.tolist()] + pivot_df.values.tolist()
            for row_idx, row_data in enumerate(pivot_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    pivot_sheet.cell(row=row_idx, column=col_idx, value=value)
            
            self.workbook.save(self.file_path)
            print(f"Pivot table created in sheet: {pivot_sheet_name}")
            return True
        except Exception as e:
            print(f"Error creating pivot table: {e}")
            return False
    
    def create_table(self, sheet_name: str, data_range: str, 
                    table_name: Optional[str] = None, 
                    table_style: str = "TableStyleMedium9") -> bool:
        """Create a native Excel table."""
        if not self.load_excel(sheet_name):
            return False
        
        try:
            # Generate table name if not provided
            if not table_name:
                table_name = f"Table_{uuid.uuid4().hex[:8]}"
            
            # Create table
            table = Table(displayName=table_name, ref=data_range)
            
            # Set table style
            style = TableStyleInfo(
                name=table_style,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            table.tableStyleInfo = style
            
            # Add table to worksheet
            self.worksheet.add_table(table)
            
            self.workbook.save(self.file_path)
            print(f"Table '{table_name}' created with range {data_range}")
            return True
        except Exception as e:
            print(f"Error creating table: {e}")
            return False
    
    def apply_formula(self, sheet_name: str, cell: str, formula: str) -> bool:
        """Apply a formula to a specific cell."""
        if not self.load_excel(sheet_name):
            return False
        
        try:
            self.worksheet[cell] = formula
            self.workbook.save(self.file_path)
            print(f"Formula '{formula}' applied to cell {cell}")
            return True
        except Exception as e:
            print(f"Error applying formula: {e}")
            return False
    
    def copy_worksheet(self, source_sheet: str, target_sheet: str) -> bool:
        """Copy a worksheet within the workbook."""
        if not self.load_excel():
            return False
        
        try:
            if source_sheet not in self.workbook.sheetnames:
                print(f"Source sheet '{source_sheet}' not found")
                return False
            
            source_ws = self.workbook[source_sheet]
            target_ws = self.workbook.copy_worksheet(source_ws)
            target_ws.title = target_sheet
            
            self.workbook.save(self.file_path)
            print(f"Sheet '{source_sheet}' copied to '{target_sheet}'")
            return True
        except Exception as e:
            print(f"Error copying worksheet: {e}")
            return False
    
    def delete_worksheet(self, sheet_name: str) -> bool:
        """Delete a worksheet from the workbook."""
        if not self.load_excel():
            return False
        
        try:
            if sheet_name not in self.workbook.sheetnames:
                print(f"Sheet '{sheet_name}' not found")
                return False
            
            if len(self.workbook.sheetnames) == 1:
                print("Cannot delete the only sheet in the workbook")
                return False
            
            self.workbook.remove(self.workbook[sheet_name])
            self.workbook.save(self.file_path)
            print(f"Sheet '{sheet_name}' deleted")
            return True
        except Exception as e:
            print(f"Error deleting worksheet: {e}")
            return False


def main():
    """Main function for command-line interface."""
    parser = argparse.ArgumentParser(
        description="Advanced Excel CLI Processor",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:

Create a new workbook:
  python excel_advanced_processor.py --create-workbook output.xlsx

Get workbook metadata:
  python excel_advanced_processor.py input.xlsx --metadata

Read data from a range:
  python excel_advanced_processor.py input.xlsx --read-data Sheet1 A1:D10

Write data to a range:
  python excel_advanced_processor.py input.xlsx --write-data Sheet1 A1 --input-data data.json

Format a range:
  python excel_advanced_processor.py input.xlsx --format-range Sheet1 A1:D1 --bold --bg-color FF0000

Create a chart:
  python excel_advanced_processor.py input.xlsx --create-chart Sheet1 A1:D10 line E1 --title "My Chart"

Create a pivot table:
  python excel_advanced_processor.py input.xlsx --create-pivot Sheet1 A1:D100 --rows Name --values Sales

Create a table:
  python excel_advanced_processor.py input.xlsx --create-table Sheet1 A1:D10 --table-name MyTable
        """
    )
    
    parser.add_argument("input_file", nargs="?", help="Path to the Excel file")
    
    # Workbook operations
    parser.add_argument("--create-workbook", help="Create a new workbook at specified path")
    parser.add_argument("--metadata", action="store_true", help="Get workbook metadata")
    parser.add_argument("--include-ranges", action="store_true", help="Include range info in metadata")
    
    # Data operations
    parser.add_argument("--read-data", nargs=2, metavar=("SHEET", "RANGE"), 
                       help="Read data from sheet and range")
    parser.add_argument("--write-data", nargs=2, metavar=("SHEET", "CELL"), 
                       help="Write data to sheet starting at cell")
    parser.add_argument("--input-data", help="JSON file containing data to write")
    parser.add_argument("--preview-only", action="store_true", help="Preview only (first 10 rows)")
    
    # Formatting operations
    parser.add_argument("--format-range", nargs=3, metavar=("SHEET", "START", "END"), 
                       help="Format a range of cells")
    parser.add_argument("--bold", action="store_true", help="Apply bold formatting")
    parser.add_argument("--italic", action="store_true", help="Apply italic formatting")
    parser.add_argument("--font-size", type=int, help="Font size")
    parser.add_argument("--font-color", help="Font color (hex)")
    parser.add_argument("--bg-color", help="Background color (hex)")
    parser.add_argument("--border-style", help="Border style")
    
    # Chart operations
    parser.add_argument("--create-chart", nargs=4, metavar=("SHEET", "RANGE", "TYPE", "TARGET"), 
                       help="Create chart (sheet, data_range, chart_type, target_cell)")
    parser.add_argument("--title", help="Chart title")
    parser.add_argument("--x-axis", help="X-axis label")
    parser.add_argument("--y-axis", help="Y-axis label")
    
    # Pivot table operations
    parser.add_argument("--create-pivot", nargs=2, metavar=("SHEET", "RANGE"), 
                       help="Create pivot table")
    parser.add_argument("--rows", nargs="+", help="Row fields for pivot table")
    parser.add_argument("--values", nargs="+", help="Value fields for pivot table")
    parser.add_argument("--columns", nargs="+", help="Column fields for pivot table")
    parser.add_argument("--agg-func", default="sum", help="Aggregation function")
    
    # Table operations
    parser.add_argument("--create-table", nargs=2, metavar=("SHEET", "RANGE"), 
                       help="Create Excel table")
    parser.add_argument("--table-name", help="Table name")
    parser.add_argument("--table-style", default="TableStyleMedium9", help="Table style")
    
    # Formula operations
    parser.add_argument("--apply-formula", nargs=3, metavar=("SHEET", "CELL", "FORMULA"), 
                       help="Apply formula to cell")
    
    # Worksheet operations
    parser.add_argument("--copy-sheet", nargs=2, metavar=("SOURCE", "TARGET"), 
                       help="Copy worksheet")
    parser.add_argument("--delete-sheet", help="Delete worksheet")
    
    # Output options
    parser.add_argument("-o", "--output", help="Output file path (for data operations)")
    
    args = parser.parse_args()
    
    # Handle workbook creation
    if args.create_workbook:
        processor = AdvancedExcelProcessor("")
        if processor.create_workbook(args.create_workbook):
            print(f"Workbook created: {args.create_workbook}")
        return
    
    # Validate input file for other operations
    if not args.input_file:
        print("Error: Input file is required for most operations")
        parser.print_help()
        return
    
    if not os.path.exists(args.input_file):
        print(f"Error: Input file '{args.input_file}' does not exist")
        return
    
    processor = AdvancedExcelProcessor(args.input_file)
    
    # Handle metadata
    if args.metadata:
        metadata = processor.get_workbook_metadata(args.include_ranges)
        print(json.dumps(metadata, indent=2))
        return
    
    # Handle data operations
    if args.read_data:
        sheet, range_str = args.read_data
        data = processor.read_data_range(sheet, range_str.split(':')[0], 
                                        range_str.split(':')[1] if ':' in range_str else None,
                                        args.preview_only)
        if args.output:
            with open(args.output, 'w') as f:
                json.dump(data, f, indent=2, default=str)
            print(f"Data saved to: {args.output}")
        else:
            print(json.dumps(data, indent=2, default=str))
        return
    
    if args.write_data:
        sheet, cell = args.write_data
        if not args.input_data:
            print("Error: --input-data is required for write operations")
            return
        
        with open(args.input_data, 'r') as f:
            data = json.load(f)
        
        processor.write_data_to_range(sheet, data, cell)
        return
    
    # Handle formatting
    if args.format_range:
        sheet, start, end = args.format_range
        processor.format_range(sheet, start, end, args.bold, args.italic,
                             args.font_size, args.font_color, args.bg_color, args.border_style)
        return
    
    # Handle chart creation
    if args.create_chart:
        sheet, range_str, chart_type, target = args.create_chart
        processor.create_chart(sheet, range_str, chart_type, target,
                             args.title or "", args.x_axis or "", args.y_axis or "")
        return
    
    # Handle pivot table creation
    if args.create_pivot:
        sheet, range_str = args.create_pivot
        if not args.rows or not args.values:
            print("Error: --rows and --values are required for pivot tables")
            return
        
        processor.create_pivot_table(sheet, range_str, args.rows, args.values,
                                   args.columns, args.agg_func)
        return
    
    # Handle table creation
    if args.create_table:
        sheet, range_str = args.create_table
        processor.create_table(sheet, range_str, args.table_name, args.table_style)
        return
    
    # Handle formula application
    if args.apply_formula:
        sheet, cell, formula = args.apply_formula
        processor.apply_formula(sheet, cell, formula)
        return
    
    # Handle worksheet operations
    if args.copy_sheet:
        source, target = args.copy_sheet
        processor.copy_worksheet(source, target)
        return
    
    if args.delete_sheet:
        processor.delete_worksheet(args.delete_sheet)
        return
    
    # If no operation specified, show help
    parser.print_help()


if __name__ == "__main__":
    main()
