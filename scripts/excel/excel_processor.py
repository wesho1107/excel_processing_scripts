#!/usr/bin/env python3
"""
Excel to Markdown Converter - AI-Assisted Workflow

A command-line tool designed to work with AI assistants for analyzing and converting Excel files.
Supports a two-step workflow:
1. Convert entire Excel sheets to markdown files for AI analysis
2. Filter and extract specific data based on AI-identified parameters
"""

import argparse
import sys
import os
from pathlib import Path
import pandas as pd
import openpyxl
from typing import List, Optional, Tuple


class ExcelProcessor:
    """Process Excel files and convert data tables to Markdown format."""
    
    def __init__(self, file_path: str):
        """Initialize the Excel processor with a file path."""
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
    
    def load_excel(self, sheet_name: Optional[str] = None) -> bool:
        """Load Excel file and select worksheet."""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            if sheet_name:
                if sheet_name in self.workbook.sheetnames:
                    self.worksheet = self.workbook[sheet_name]
                else:
                    print(f"Warning: Sheet '{sheet_name}' not found. Using first sheet.")
                    self.worksheet = self.workbook.active
            else:
                self.worksheet = self.workbook.active
            return True
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return False
    
    def convert_all_sheets_to_markdown(self, output_dir: str = "sheets_markdown") -> List[str]:
        """
        Convert all sheets in the Excel file to markdown files for AI analysis.
        This is Step 2 of the AI-assisted workflow.
        
        Args:
            output_dir: Directory to save markdown files
            
        Returns:
            List of generated file paths
        """
        if not self.load_excel():
            return []
        
        os.makedirs(output_dir, exist_ok=True)
        generated_files = []
        
        for sheet_name in self.workbook.sheetnames:
            print(f"Processing sheet: {sheet_name}")
            
            worksheet = self.workbook[sheet_name]
            markdown_content = self._sheet_to_markdown(worksheet, sheet_name)
            
            # Create safe filename
            safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            output_file = os.path.join(output_dir, f"{safe_name}.md")
            
            try:
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(markdown_content)
                print(f"Saved {sheet_name} to {output_file}")
                generated_files.append(output_file)
            except Exception as e:
                print(f"Error saving {sheet_name}: {e}")
        
        return generated_files
    
    def _sheet_to_markdown(self, worksheet, sheet_name: str) -> str:
        """Convert an entire worksheet to markdown format with cell references for AI analysis."""
        lines = [f"# Sheet: {sheet_name}\n", "## Raw Data with Cell References\n"]
        
        # Collect all data with cell references
        data_rows = []
        for row in range(1, worksheet.max_row + 1):
            row_data = []
            has_content = False
            
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                
                if cell_value is not None:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    row_data.append(f"{cell_value} ({col_letter}{row})")
                    has_content = True
                else:
                    row_data.append("")
            
            if has_content:
                data_rows.append(row_data)
        
        # Convert to DataFrame and generate markdown table
        if data_rows:
            max_cols = max(len(row) for row in data_rows)
            padded_rows = [row + [""] * (max_cols - len(row)) for row in data_rows]
            
            df = pd.DataFrame(padded_rows, columns=[f"Col {i+1}" for i in range(max_cols)])
            lines.append(df.to_markdown(index=False, tablefmt='github'))
        
        # Add summary
        lines.extend([
            "\n## Summary for AI Analysis",
            f"- Total rows with data: {len(data_rows)}",
            f"- Total columns: {worksheet.max_column}",
            f"- Sheet name: {sheet_name}",
            "- Cell references are included in parentheses for precise identification"
        ])
        
        return "\n".join(lines)
    
    def filter_data_by_criteria(self, start_cell: str, columns: Optional[List[str]], 
                               primary_column: str, key_value: str, 
                               output_path: str, sheet_name: Optional[str] = None) -> str:
        """
        Filter data based on AI-identified criteria.
        This is Step 5 of the AI-assisted workflow.
        
        Args:
            start_cell: Cell reference where data table starts (e.g., "A6")
            columns: List of column names to extract (None = all columns)
            primary_column: Column to filter by
            key_value: Value to filter for in the primary column
            output_path: Path to save filtered markdown
            sheet_name: Sheet name to process (optional)
            
        Returns:
            Generated markdown content or empty string if failed
        """
        if not self.load_excel(sheet_name):
            return ""
        
        start_row, start_col = self._parse_cell_reference(start_cell)
        print(f"Starting data extraction from cell {start_cell} (row {start_row}, col {start_col})")
        
        # Extract and validate headers
        headers = self._extract_headers_from_position(start_row, start_col)
        print(f"Found headers: {headers}")
        
        # If no columns specified, use all available columns
        if columns is None:
            columns = [h for h in headers if h]  # Filter out empty headers
            print(f"No columns specified, using all columns: {columns}")
        
        column_indices = self._find_column_indices(headers, columns)
        primary_col_index = self._find_column_indices(headers, [primary_column])
        
        if not column_indices:
            print("Error: None of the specified columns were found.")
            return ""
        
        if not primary_col_index:
            print(f"Error: Primary column '{primary_column}' not found.")
            return ""
        
        # Extract and filter data
        filtered_data = self._extract_filtered_data(
            start_row + 1, start_col, column_indices, 
            primary_col_index[0], key_value
        )
        
        if not filtered_data:
            print(f"No data found for {primary_column} = '{key_value}'")
            return ""
        
        # Create and save markdown
        filtered_headers = [headers[i] for i in column_indices]
        df = pd.DataFrame(filtered_data, columns=filtered_headers)
        
        print(f"Filtered data: {len(df)} rows matching criteria")
        
        markdown_content = self._create_filtered_markdown(df, primary_column, key_value, start_cell)
        
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(markdown_content)
            print(f"Filtered data saved to: {output_path}")
        except Exception as e:
            print(f"Error saving file: {e}")
            return ""
        
        return markdown_content
    
    def _parse_cell_reference(self, cell_ref: str) -> Tuple[int, int]:
        """Parse cell reference like 'A3' into row and column numbers."""
        col_str = ""
        row_str = ""
        
        for char in cell_ref:
            if char.isalpha():
                col_str += char
            else:
                row_str += char
        
        col_num = openpyxl.utils.column_index_from_string(col_str)
        row_num = int(row_str)
        
        return row_num, col_num
    
    def _extract_headers_from_position(self, row: int, col: int) -> List[str]:
        """Extract headers from a specific position."""
        headers = []
        max_col = self.worksheet.max_column
        
        for c in range(col, max_col + 1):
            cell_value = self.worksheet.cell(row=row, column=c).value
            if cell_value is not None:
                headers.append(str(cell_value).strip())
            else:
                if headers:  # Stop at first empty cell after finding headers
                    break
                headers.append("")
        
        return headers
    
    def _find_column_indices(self, headers: List[str], target_columns: List[str]) -> List[int]:
        """Find indices of target columns in headers."""
        indices = []
        for target in target_columns:
            for i, header in enumerate(headers):
                if header.lower() == target.lower():
                    indices.append(i)
                    break
        return indices
    
    def _extract_filtered_data(self, start_row: int, start_col: int, 
                              column_indices: List[int], primary_col_index: int, 
                              key_value: str) -> List[List[str]]:
        """Extract data rows that match the filter criteria."""
        filtered_rows = []
        max_row = self.worksheet.max_row
        
        for row in range(start_row, max_row + 1):
            # Check if this row matches the filter criteria
            primary_cell_value = self.worksheet.cell(row=row, column=start_col + primary_col_index).value
            
            if primary_cell_value is not None and str(primary_cell_value).strip().lower() == key_value.lower():
                # Extract data for specified columns
                row_data = []
                for col_idx in column_indices:
                    cell_value = self.worksheet.cell(row=row, column=start_col + col_idx).value
                    row_data.append(str(cell_value).strip() if cell_value is not None else "")
                
                filtered_rows.append(row_data)
        
        return filtered_rows
    
    def _create_filtered_markdown(self, df: pd.DataFrame, primary_column: str, 
                                 key_value: str, start_cell: str) -> str:
        """Create markdown content with metadata for filtered data."""
        lines = [
            f"# Filtered Data: {primary_column} = '{key_value}'",
            "",
            "## Metadata",
            f"- Filter criteria: {primary_column} = '{key_value}'",
            f"- Data source starting cell: {start_cell}",
            f"- Total matching rows: {len(df)}",
            f"- Columns included: {', '.join(df.columns)}",
            "",
            "## Data Table",
            ""
        ]
        
        # Add the actual data table
        lines.append(df.to_markdown(index=False))
        
        return "\n".join(lines)


def main():
    """Main function for command-line interface with AI-assisted workflow support."""
    parser = argparse.ArgumentParser(
        description="Excel to Markdown Converter - AI-Assisted Workflow",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
AI-Assisted Workflow Examples:

Step 2 - Convert all sheets to markdown for AI analysis:
  python excel_processor.py input.xlsx --convert-all

Step 5 - Filter data based on AI analysis:
  python excel_processor.py input.xlsx -s "A6" -c "Name" "Job" -pc "Department" -k "Engineering" -o "engineering.md"
        """
    )
    
    parser.add_argument("input_file", help="Path to the Excel file")
    
    # AI-assisted workflow options
    parser.add_argument("--convert-all", action="store_true",
                       help="Convert all sheets to markdown files for AI analysis (Step 2)")
    parser.add_argument("-s", "--start-cell", 
                       help="Starting cell of data table (e.g., 'A6') for filtering (Step 5)")
    parser.add_argument("-c", "--columns", nargs="*",
                       help="List of column names to extract (Step 5). If not specified, all columns will be included.")
    parser.add_argument("-pc", "--primary-column", 
                       help="Primary column to filter by (Step 5)")
    parser.add_argument("-k", "--key-value", 
                       help="Value to filter for in primary column (Step 5)")
    parser.add_argument("-o", "--output", help="Output file path")
    parser.add_argument("--sheet", help="Sheet name to process")
    
    args = parser.parse_args()
    
    # Validate input file
    if not os.path.exists(args.input_file):
        print(f"Error: Input file '{args.input_file}' does not exist.")
        sys.exit(1)
    
    processor = ExcelProcessor(args.input_file)
    
    # Step 2: Convert all sheets to markdown
    if args.convert_all:
        print("Converting all sheets to markdown for AI analysis...")
        output_dir = args.output or "sheets_markdown"
        generated_files = processor.convert_all_sheets_to_markdown(output_dir)
        
        if generated_files:
            print(f"\nGenerated {len(generated_files)} markdown files:")
            for file in generated_files:
                print(f"  - {file}")
            print(f"\nNext: Have the AI analyze these files to identify:")
            print("  - Column headers")
            print("  - Starting cell of actual data")
            print("  - Primary keys and data structure")
        else:
            print("Failed to generate markdown files.")
            sys.exit(1)
        return
    
    # Step 5: Filter data based on AI analysis
    if args.start_cell and args.primary_column and args.key_value:
        if not args.output:
            print("Error: Output file (-o) is required for filtering.")
            sys.exit(1)
        
        print("Filtering data based on AI analysis...")
        result = processor.filter_data_by_criteria(
            args.start_cell, args.columns, args.primary_column, 
            args.key_value, args.output, args.sheet
        )
        
        if not result:
            print("Failed to filter data.")
            sys.exit(1)
        return
    
    # Show usage if no valid mode specified
    print("Error: Please specify either --convert-all or provide filtering parameters.")
    print("Use --help for usage examples.")
    sys.exit(1)


if __name__ == "__main__":
    main()